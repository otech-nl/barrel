try:
    from openpyxl import load_workbook
except ImportError as e:
    module = str(e).split()[-1]
    print('Please run "pip install %s"' % module)


class ExcelObject(object):
    ''' an object that is derived from an Excel spreadsheet '''

    key_column = 'id'
    field_names = []

    def __init__(self, row):
        self.values = dict()
        ''' derive an object from a worksheet row '''
        for i in range(0, len(self.field_names)):
            self.values[self.field_names[i]] = row[i].value
        self.id = self.values[self.key_column]

    @classmethod
    def sheet2array(cls, sheet, first_row=1):
        ''' derive objects from a worksheet '''
        print('Reading %s records (%s) from row %d with key column "%s"' %
              (cls.__name__, sheet, first_row, cls.key_column))
        cls.field_names = []
        key_column = None
        records = []
        for row in sheet.iter_rows():
            row_nr = int(row[0].row)
            if row_nr < first_row:
                # print '   ignoring line %d: %s' % (row[0].row, row[0].value)
                continue
            elif row_nr == first_row:
                # print 'HEADERS'
                for field in row:
                    if not field.value:
                        break
                    cls.field_names.append(field.value)
                key_column = cls.field_names.index(cls.key_column)
                if not key_column:
                    raise KeyError('Key column %s for %s not found' %
                                   (cls.key_column, cls.__name__))
            else:
                # print 'RECORD %d: %s' % (key_column, cls.field_names)
                if row[key_column].value:
                    record = cls(row)
                    records.append(record)
        return records

    @staticmethod
    def load(filename):
        return load_workbook(filename)

    def __repr__(self):
        return '%s@%s' % (self.__class__.__name__, self.id)
