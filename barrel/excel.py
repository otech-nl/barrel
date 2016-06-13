from openpyxl import load_workbook

class ExcelObject(object):
    ''' an object that is derived from an Excel spreadsheet '''

    key_column = 'id'
    field_names = []

    def __init__(self, row):
        self.values = dict()
        ''' derive an object from a worksheet row '''
        for i in range(0, len(self.field_names)):
            self.values[self.field_names[i]] = row[i].value
        self.id = self.values[self.key_column]

    @classmethod
    def sheet2array(cls, sheet, first_row=1):
        ''' derive objects from a worksheet '''
        # print 'Reading %s records (%s)' % (cls.__name__, sheet)
        key_column = None
        records = []
        for row in sheet.iter_rows():
            if int(row[0].row) < first_row:
                # print '   ignoring line %d: %s' % (row[0].row, row[0].value)
                continue
            if cls.field_names:
                # print 'RECORD %d' % key_column
                if row[key_column].value:
                    record = cls(row)
                    records.append(record)
            else:
                # print 'HEADERS'
                for field in row:
                    if not field.value: break
                    cls.field_names.append(field.value)
                key_column = cls.field_names.index(cls.key_column)
        return records

    @staticmethod
    def load(filename):
        return load_workbook(filename)

    def __repr__(self):
        return '%s@%s' % (self.__class__.__name__, self.id)
